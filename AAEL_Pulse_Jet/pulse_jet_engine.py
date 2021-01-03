from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import scipy

importData = False
plotGraphs = True
# frequencyOfSampling, inletTemperatureToV, pressureToV, exhaustTemperatureToVo, fuelPressureToV, thrustToV,
constants = [1/3500, 120, 120, 1.6, 1, 35]
labels = ["Inlet temperature [C]",
          "Outlet temperature [C]",
          "Fuel pressure [bar]",
          "Combustion chamber pressure [bar]"]

if importData:
  # import of the data
  wb = load_workbook('Pulse engine_test1_1.xlsx', read_only=True)
  ws = wb[wb.sheetnames[0]]
  data = np.array([[i.value for i in j[0:5]] for j in ws.rows])
  names = data[0, :]
  for columns in range(data.shape[1]):
    savedArray = data[1:, columns].astype('float64')
    savedArray = np.multiply(savedArray, constants[columns])
    np.savetxt('data/'+names[columns]+'.txt', savedArray)
  np.savetxt('data/names.txt', names, fmt='%s')
  del ws, wb

#calculating the data

#
plt.rcParams['font.size'] = '8'
plt.rcParams['font.family'] = 'Times New Roman'
fig, ax = plt.subplots(4, 1)

names = np.loadtxt('data/names.txt', delimiter='\n', dtype=str)
time = np.loadtxt('data/'+names[0]+".txt")
names = names[1:]
if plotGraphs:
  for i in range(names.shape[0]):
      plottedArray = np.loadtxt('data/'+names[i]+'.txt', dtype=float)
      ax[i].plot(time, plottedArray, color='k')
      ax[i].set_title(labels[i])
  ax[-1].set_xlabel("time [s]")

pressure_file = None
for name in names:
  str.lower(name)
  if str.find(name, "chamber") > 0:
    pressure_file = name
pressure = np.loadtxt('data/'+pressure_file+'.txt', dtype=float)
pressure_peak_time = time
previous = 0
toDelet = []
change = []
new_pressure = []
new_time = []

width = 3
for i in range(round(pressure.shape[0]/width)):
  executing = pressure[i*width:i*width+width-1].tolist()
  maximum = np.max(executing)
  timeMaximum = time[i*width+executing.index(maximum)]
  new_pressure.append(maximum)
  new_time.append(timeMaximum)

peaks = []
times = []
change = 0
usedPressure = pressure
usedTime = time
for i in range(len(usedPressure)-2):
  change_one = usedPressure[i+1]-usedPressure[i]
  change_two = usedPressure[i+2]-usedPressure[i+1]
  if change_one > 0:
    if change_one*change_two <= 0:
      peaks.append(usedPressure[i+1])
      times.append(usedTime[i+1])

V_actual = scipy.fft.fft(usedPressure)
fig2, az = plt.subplots(2, 1)
az[0].plot(time, pressure, color='k')
az[0].set_title("Combustion chamber pressure")
az[0].set_ylabel("pressure [bar]")
az[0].set_xlabel("time [s]")
az[1].plot(np.abs(V_actual[:len(usedPressure) // 2]), color='k')
az[1].set_xlim(0, 500)
az[1].set_title("FFT result")
az[1].set_xlabel("frequency [Hz]")
az[0].scatter(times, peaks, color='r')
fig.tight_layout(pad=1.0)
fig2.tight_layout(pad=2.0)

fig.savefig('results/plots.png')
fig2.savefig('results/frequency.png')
plt.show()